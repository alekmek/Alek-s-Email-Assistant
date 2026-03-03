import io
import re
from pathlib import Path
from typing import Optional

from docx import Document
from openpyxl import load_workbook
from PIL import Image


class AttachmentProcessor:
    """Process email attachments for LLM analysis."""

    MAX_TEXT_CHARS = 12000

    # Supported content types
    PDF_TYPES = ["application/pdf"]
    IMAGE_TYPES = ["image/png", "image/jpeg", "image/gif", "image/webp"]
    WORD_TYPES = [
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ]
    EXCEL_TYPES = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]
    TEXT_TYPES = ["text/plain", "text/csv", "text/html"]
    IMAGE_EXT_TO_MIME = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }
    PDF_EXTENSIONS = {".pdf"}
    WORD_EXTENSIONS = {".docx", ".doc"}
    EXCEL_EXTENSIONS = {".xlsx", ".xls"}
    TEXT_EXTENSIONS = {".txt", ".csv", ".html", ".htm"}

    def _normalize_content_type(self, content_type: str) -> str:
        value = (content_type or "").strip().lower()
        if ";" in value:
            value = value.split(";", 1)[0].strip()
        return value

    def _normalize_filename(self, filename: str) -> str:
        value = (filename or "").strip()
        # Common malformed attachment names can contain spaces before the extension dot.
        return re.sub(r"\s+\.", ".", value)

    def _extension(self, filename: str) -> str:
        return Path(filename).suffix.lower()

    def _looks_like_pdf(self, data: bytes) -> bool:
        return data[:5] == b"%PDF-"

    def _truncate_text(self, text: str, max_chars: int | None = None) -> tuple[str, bool]:
        limit = max_chars or self.MAX_TEXT_CHARS
        value = (text or "").strip()
        if len(value) <= limit:
            return value, False
        return value[:limit], True

    def process(self, data: bytes, content_type: str, filename: str) -> dict:
        """
        Process attachment and return content suitable for the LLM.

        Returns a dict with either:
        - {"type": "summary", ...} for image/document metadata summaries
        - {"type": "text", "content": "..."} for extracted text
        """
        normalized_type = self._normalize_content_type(content_type)
        normalized_filename = self._normalize_filename(filename)
        extension = self._extension(normalized_filename)

        if (
            normalized_type in self.PDF_TYPES
            or extension in self.PDF_EXTENSIONS
            or self._looks_like_pdf(data)
        ):
            return self._process_pdf(data, normalized_filename)

        if normalized_type in self.IMAGE_TYPES or extension in self.IMAGE_EXT_TO_MIME:
            image_mime = normalized_type or self.IMAGE_EXT_TO_MIME.get(extension, "image/jpeg")
            return self._process_image(data, image_mime)

        if normalized_type in self.WORD_TYPES or extension in self.WORD_EXTENSIONS:
            return self._process_word(data, normalized_filename)

        if normalized_type in self.EXCEL_TYPES or extension in self.EXCEL_EXTENSIONS:
            return self._process_excel(data, normalized_filename)

        if normalized_type in self.TEXT_TYPES or extension in self.TEXT_EXTENSIONS:
            return self._process_text(data, normalized_filename)

        if normalized_type.startswith("text/"):
            return self._process_text(data, normalized_filename)

        if normalized_type.startswith("image/"):
            return self._process_image(data, normalized_type)

        else:
            return {
                "type": "unsupported",
                "message": (
                    f"Cannot process attachment of type {content_type} "
                    f"(normalized: {normalized_type}). Filename: {normalized_filename}"
                ),
            }

    def _process_pdf(self, data: bytes, filename: str) -> dict:
        """Process PDF with text extraction; otherwise return compact metadata summary."""
        page_count: int | None = None
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(data))
            page_count = len(reader.pages)
            page_text: list[str] = []
            for page in reader.pages:
                extracted = (page.extract_text() or "").strip()
                if extracted:
                    page_text.append(extracted)

            if page_text:
                combined = "\n\n".join(page_text).strip()
                combined, truncated = self._truncate_text(combined)

                return {
                    "type": "text",
                    "content": f"Content of {filename}:\n\n{combined}",
                    "extracted_from": "pdf",
                    "truncated": truncated,
                    "page_count": page_count,
                }
        except Exception:
            # Fall back to metadata summary if extraction is unavailable/failed.
            pass

        return {
            "type": "summary",
            "analysis_type": "document",
            "summary": (
                f"PDF attachment '{filename}' has no extractable text content. "
                f"Page count: {page_count if page_count is not None else 'unknown'}."
            ),
            "metadata": {
                "filename": filename,
                "page_count": page_count,
            },
        }

    def _process_image(self, data: bytes, content_type: str) -> dict:
        """Return a compact image metadata summary (without raw binary payload)."""
        width = None
        height = None
        mode = None
        image_format = None
        frames = None
        is_animated = False
        error = None

        try:
            with Image.open(io.BytesIO(data)) as image:
                width, height = image.size
                mode = image.mode
                image_format = image.format
                frames = getattr(image, "n_frames", 1)
                is_animated = bool(frames and frames > 1)
        except Exception as exc:
            error = str(exc)

        if error:
            summary = (
                "Image attachment was downloaded, but metadata extraction failed. "
                "The file may be corrupted or unsupported."
            )
        else:
            summary = (
                f"Image attachment metadata: format={image_format or 'unknown'}, "
                f"size={width}x{height}, mode={mode or 'unknown'}, "
                f"animated={'yes' if is_animated else 'no'}."
            )

        return {
            "type": "summary",
            "analysis_type": "image",
            "summary": summary,
            "metadata": {
                "content_type": content_type,
                "width": width,
                "height": height,
                "mode": mode,
                "format": image_format,
                "frames": frames,
                "animated": is_animated,
                "metadata_error": error,
            },
        }

    def _process_word(self, data: bytes, filename: str) -> dict:
        """Extract text from Word document."""
        try:
            doc = Document(io.BytesIO(data))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            text = "\n\n".join(paragraphs)
            text, truncated = self._truncate_text(text)
            return {
                "type": "text",
                "content": f"Content of {filename}:\n\n{text}",
                "truncated": truncated,
            }
        except Exception as e:
            return {
                "type": "error",
                "message": f"Failed to process Word document {filename}: {str(e)}",
            }

    def _process_excel(self, data: bytes, filename: str) -> dict:
        """Extract data from Excel spreadsheet."""
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True)
            result_parts = [f"Content of {filename}:"]

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                result_parts.append(f"\n## Sheet: {sheet_name}\n")

                rows = []
                for row in sheet.iter_rows(max_row=100, values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        row_text = " | ".join(
                            str(cell) if cell is not None else "" for cell in row
                        )
                        rows.append(row_text)

                result_parts.append("\n".join(rows))

            text, truncated = self._truncate_text("\n".join(result_parts))
            return {
                "type": "text",
                "content": text,
                "truncated": truncated,
            }
        except Exception as e:
            return {
                "type": "error",
                "message": f"Failed to process Excel file {filename}: {str(e)}",
            }

    def _process_text(self, data: bytes, filename: str) -> dict:
        """Process plain text files."""
        try:
            text = data.decode("utf-8")
            text, truncated = self._truncate_text(text)
            return {
                "type": "text",
                "content": f"Content of {filename}:\n\n{text}",
                "truncated": truncated,
            }
        except UnicodeDecodeError:
            try:
                text = data.decode("latin-1")
                text, truncated = self._truncate_text(text)
                return {
                    "type": "text",
                    "content": f"Content of {filename}:\n\n{text}",
                    "truncated": truncated,
                }
            except Exception as e:
                return {
                    "type": "error",
                    "message": f"Failed to decode text file {filename}: {str(e)}",
                }


# Singleton instance
_processor: Optional[AttachmentProcessor] = None


def get_attachment_processor() -> AttachmentProcessor:
    global _processor
    if _processor is None:
        _processor = AttachmentProcessor()
    return _processor
